# call_center_project/app/rh/routes.py

from flask import Blueprint, render_template, request, flash, redirect, url_for, current_app, send_from_directory, Response
from flask_login import login_required, current_user
from datetime import datetime
from app import db
from app.models_rh import Funcionario, Cargo, Departamento, DocumentoFuncionario, Afastamento
from app.decorators import require_plan
from .validators import is_cpf_valid
from .calculos import calcular_rescisao, calcular_folha_pagamento, _get_dias_uteis_no_mes
import os
from werkzeug.utils import secure_filename
import openpyxl
from io import BytesIO
import json

rh = Blueprint('rh', __name__, url_prefix='/rh')
def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf', 'xlsx'}
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@rh.route('/dashboard')
@login_required
@require_plan('medio')
def dashboard():
    empresa_id = current_user.empresa_id
    total_funcionarios = Funcionario.query.filter_by(empresa_id=empresa_id, status='ativo').count()
    total_departamentos = Departamento.query.filter_by(empresa_id=empresa_id).count()
    total_cargos = Cargo.query.filter_by(empresa_id=empresa_id).count()
    
    return render_template('rh/dashboard.html', 
        total_funcionarios=total_funcionarios,
        total_departamentos=total_departamentos,
        total_cargos=total_cargos)

@rh.route('/dashboard_financeiro')
@login_required
@require_plan('medio')
def dashboard_financeiro():
    return render_template('rh/dashboard_financeiro.html')

@rh.route('/funcionarios')
@login_required
@require_plan('medio')
def listar_funcionarios():
    empresa_id = current_user.empresa_id
    funcionarios = Funcionario.query.filter_by(empresa_id=empresa_id).order_by(Funcionario.nome).all()
    return render_template('rh/funcionarios.html', funcionarios=funcionarios)

def _populate_funcionario_from_form(funcionario, form_data):
    """Função auxiliar para popular dados do funcionário a partir do formulário."""
    def to_float(value):
        if not isinstance(value, str):
            return 0.0
        # Remove R$, espaços e troca o separador de milhar por nada
        clean_value = value.replace('R$', '').strip().replace('.', '')
        # Troca a vírgula do decimal por ponto
        return float(clean_value.replace(',', '.') or 0)

    funcionario.nome = form_data.get('nome')
    funcionario.cpf = form_data.get('cpf')
    funcionario.rg = form_data.get('rg')
    data_nascimento_str = form_data.get('data_nascimento')
    if data_nascimento_str:
        funcionario.data_nascimento = datetime.strptime(data_nascimento_str, '%Y-%m-%d').date()
    
    funcionario.sexo = form_data.get('sexo')
    funcionario.estado_civil = form_data.get('estado_civil')
    funcionario.telefone = form_data.get('telefone')
    funcionario.email = form_data.get('email')
    funcionario.endereco = form_data.get('endereco')
    funcionario.cep = form_data.get('cep')
    funcionario.cidade = form_data.get('cidade')
    funcionario.estado = form_data.get('estado')
    funcionario.cargo_id = form_data.get('cargo_id')
    funcionario.departamento_id = form_data.get('departamento_id')
    funcionario.salario = to_float(form_data.get('salario'))
    
    data_admissao_str = form_data.get('data_admissao')
    if data_admissao_str:
        funcionario.data_admissao = datetime.strptime(data_admissao_str, '%Y-%m-%d').date()

    funcionario.jornada_trabalho = form_data.get('jornada_trabalho')
    
    funcionario.recebe_vt = 'recebe_vt' in form_data
    funcionario.vale_transporte_diario = to_float(form_data.get('vale_transporte_diario'))
    
    funcionario.recebe_va = 'recebe_va' in form_data
    funcionario.vale_alimentacao_diario = to_float(form_data.get('vale_alimentacao_diario'))

    funcionario.recebe_vr = 'recebe_vr' in form_data
    funcionario.vale_refeicao_diario = to_float(form_data.get('vale_refeicao_diario'))

@rh.route('/funcionarios/novo', methods=['GET', 'POST'])
@login_required
@require_plan('medio')
def novo_funcionario():
    if request.method == 'POST':
        cpf = request.form['cpf']
        if not is_cpf_valid(cpf):
            flash('O CPF inserido é inválido. Por favor, verifique.', 'danger')
            cargos = Cargo.query.filter_by(empresa_id=current_user.empresa_id).all()
            departamentos = Departamento.query.filter_by(empresa_id=current_user.empresa_id).all()
            return render_template('rh/funcionario_form.html', cargos=cargos, departamentos=departamentos, form_data=request.form)

        try:
            novo_func = Funcionario(empresa_id=current_user.empresa_id)
            _populate_funcionario_from_form(novo_func, request.form)

            ultimo = Funcionario.query.order_by(Funcionario.id.desc()).first()
            novo_func.matricula = f"FUNC{str(ultimo.id + 1 if ultimo else 1).zfill(4)}"
            
            db.session.add(novo_func)
            db.session.commit()
            flash('Funcionário cadastrado com sucesso!', 'success')
            return redirect(url_for('rh.listar_funcionarios'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar funcionário: {str(e)}', 'danger')

    cargos = Cargo.query.filter_by(empresa_id=current_user.empresa_id).all()
    departamentos = Departamento.query.filter_by(empresa_id=current_user.empresa_id).all()
    return render_template('rh/funcionario_form.html', cargos=cargos, departamentos=departamentos)

@rh.route('/funcionarios/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@require_plan('medio')
def editar_funcionario(id):
    funcionario = Funcionario.query.get_or_404(id)
    if funcionario.empresa_id != current_user.empresa_id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('rh.listar_funcionarios'))

    if request.method == 'POST':
        try:
            _populate_funcionario_from_form(funcionario, request.form)
            db.session.commit()
            flash('Funcionário atualizado com sucesso!', 'success')
            return redirect(url_for('rh.ver_funcionario', id=id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar funcionário: {str(e)}', 'danger')
    
    cargos = Cargo.query.filter_by(empresa_id=current_user.empresa_id).order_by(Cargo.nome).all()
    departamentos = Departamento.query.filter_by(empresa_id=current_user.empresa_id).order_by(Departamento.nome).all()
    return render_template('rh/funcionario_form.html', funcionario=funcionario, cargos=cargos, departamentos=departamentos)

@rh.route('/funcionarios/<int:id>')
@login_required
@require_plan('medio')
def ver_funcionario(id):
    funcionario = Funcionario.query.get_or_404(id)
    if funcionario.empresa_id != current_user.empresa_id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('rh.listar_funcionarios'))

    documentos_pessoais = [doc for doc in funcionario.documentos if doc.tipo_documento != 'Atestado']
    atestados = [doc for doc in funcionario.documentos if doc.tipo_documento == 'Atestado']
    afastamentos = funcionario.afastamentos

    hoje = datetime.today()
    dias_uteis = _get_dias_uteis_no_mes(hoje.year, hoje.month, funcionario.jornada_trabalho)
    
    total_vt = dias_uteis * (funcionario.vale_transporte_diario or 0)
    total_va = dias_uteis * (funcionario.vale_alimentacao_diario or 0)
    total_vr = dias_uteis * (funcionario.vale_refeicao_diario or 0)

    return render_template('rh/ver_funcionario.html', 
        funcionario=funcionario, 
        documentos_pessoais=documentos_pessoais,
        atestados=atestados,
        afastamentos=afastamentos,
        dias_uteis=dias_uteis,
        total_vt=total_vt,
        total_va=total_va,
        total_vr=total_vr)

@rh.route('/funcionarios/<int:id>/upload', methods=['POST'])
@login_required
@require_plan('medio')
def upload_documento(id):
    # (Este código não foi alterado)
    funcionario = Funcionario.query.get_or_404(id)
    if funcionario.empresa_id != current_user.empresa_id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('rh.listar_funcionarios'))
    
    upload_type = request.form.get('upload_type')
    
    if upload_type not in ['foto', 'documento', 'atestado']:
        flash('Tipo de upload inválido.', 'danger')
        return redirect(url_for('rh.ver_funcionario', id=id))

    if 'file' not in request.files:
        flash('Nenhum ficheiro selecionado.', 'warning')
        return redirect(url_for('rh.ver_funcionario', id=id))

    file = request.files['file']
    if file.filename == '':
        flash('Nenhum ficheiro selecionado.', 'warning')
        return redirect(url_for('rh.ver_funcionario', id=id))

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        upload_path = os.path.join(current_app.instance_path, 'uploads', str(current_user.empresa_id), str(id))
        os.makedirs(upload_path, exist_ok=True)
        file_path = os.path.join(upload_path, filename)
        file.save(file_path)

        if upload_type == 'foto':
            funcionario.foto_perfil = filename
            flash('Foto de perfil atualizada com sucesso!', 'success')
        else:
            tipo_doc = 'Atestado' if upload_type == 'atestado' else request.form.get('tipo_documento', 'Outro')
            novo_doc = DocumentoFuncionario(
                nome_arquivo=filename,
                tipo_documento=tipo_doc,
                caminho_arquivo=file_path,
                funcionario_id=id
            )
            db.session.add(novo_doc)
            flash(f'{tipo_doc} enviado com sucesso!', 'success')
        
        db.session.commit()
    else:
        flash('Tipo de ficheiro não permitido. Apenas: png, jpg, jpeg, pdf.', 'danger')

    return redirect(url_for('rh.ver_funcionario', id=id))

@rh.route('/uploads/<int:empresa_id>/<int:funcionario_id>/<filename>')
@login_required
def uploaded_file(empresa_id, funcionario_id, filename):
    # (Este código não foi alterado)
    if empresa_id != current_user.empresa_id:
        return "Acesso negado", 403
    path = os.path.join(current_app.instance_path, 'uploads', str(empresa_id), str(funcionario_id))
    return send_from_directory(path, filename)

@rh.route('/funcionarios/<int:id>/rescisao', methods=['GET', 'POST'])
@login_required
@require_plan('medio')
def calcular_rescisao_funcionario(id):
    # (Este código não foi alterado)
    funcionario = Funcionario.query.get_or_404(id)
    if funcionario.empresa_id != current_user.empresa_id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('rh.listar_funcionarios'))

    resultado = None
    if request.method == 'POST':
        data_demissao = datetime.strptime(request.form['data_demissao'], '%Y-%m-%d').date()
        motivo = request.form['motivo']
        aviso_indenizado = 'aviso_indenizado' in request.form
        ferias_vencidas = 'ferias_vencidas' in request.form
        
        resultado = calcular_rescisao(
            salario_bruto=funcionario.salario,
            data_admissao=funcionario.data_admissao,
            data_demissao=data_demissao,
            motivo=motivo,
            aviso_previo_indenizado=aviso_indenizado,
            ferias_vencidas=ferias_vencidas
        )

    return render_template('rh/rescisao_form.html', funcionario=funcionario, resultado=resultado)

@rh.route('/funcionarios/<int:id>/folha_pagamento', methods=['GET', 'POST'])
@login_required
@require_plan('medio')
def simular_folha_pagamento(id):
    # (Este código não foi alterado)
    funcionario = Funcionario.query.get_or_404(id)
    if funcionario.empresa_id != current_user.empresa_id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('rh.listar_funcionarios'))

    resultado = None
    mes, ano = datetime.today().month, datetime.today().year

    if request.method == 'POST':
        mes = int(request.form.get('mes', mes))
        ano = int(request.form.get('ano', ano))

    resultado = calcular_folha_pagamento(funcionario, ano=ano, mes=mes)

    return render_template('rh/folha_pagamento_form.html', 
        funcionario=funcionario, 
        resultado=resultado,
        mes_selecionado=mes,
        ano_selecionado=ano)


@rh.route('/funcionarios/importar', methods=['POST'])
@login_required
@require_plan('medio')
def importar_funcionarios():
    # (Este código não foi alterado)
    if 'excel_file' not in request.files:
        flash('Nenhum ficheiro selecionado.', 'warning')
        return redirect(url_for('rh.listar_funcionarios'))

    file = request.files['excel_file']

    if file.filename == '' or not allowed_file(file.filename):
        flash('Ficheiro inválido ou não selecionado. Por favor, use um ficheiro .xlsx.', 'danger')
        return redirect(url_for('rh.listar_funcionarios'))

    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        
        header = [cell.value for cell in sheet[1]]
        
        sucesso = 0
        erros = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            data = dict(zip(header, row))
            
            if not all([data.get('nome'), data.get('cpf'), data.get('email'), data.get('cargo'), data.get('departamento')]):
                erros += 1
                continue
            
            if not is_cpf_valid(data['cpf']):
                erros += 1
                continue

            cargo = Cargo.query.filter_by(nome=data['cargo'], empresa_id=current_user.empresa_id).first()
            departamento = Departamento.query.filter_by(nome=data['departamento'], empresa_id=current_user.empresa_id).first()

            if not cargo or not departamento:
                erros += 1
                continue
            
            try:
                data_nasc = datetime.strptime(str(data['data_nascimento']), '%Y-%m-%d %H:%M:%S').date()
                data_adm = datetime.strptime(str(data['data_admissao']), '%Y-%m-%d %H:%M:%S').date()
            except (ValueError, TypeError):
                erros += 1
                continue

            novo_func = Funcionario(
                nome=data['nome'],
                cpf=data['cpf'],
                rg=data.get('rg', ''),
                data_nascimento=data_nasc,
                sexo=data.get('sexo', 'M'),
                estado_civil=data.get('estado_civil', 'Solteiro(a)'),
                telefone=data.get('telefone', ''),
                email=data['email'],
                endereco=data.get('endereco', ''),
                cep=data.get('cep', ''),
                cidade=data.get('cidade', ''),
                estado=data.get('estado', ''),
                salario=float(data.get('salario', 0)),
                data_admissao=data_adm,
                cargo_id=cargo.id,
                departamento_id=departamento.id,
                empresa_id=current_user.empresa_id
            )
            ultimo = Funcionario.query.order_by(Funcionario.id.desc()).first()
            novo_func.matricula = f"FUNC{str(ultimo.id + 1 if ultimo else 1).zfill(4)}"
            
            db.session.add(novo_func)
            sucesso += 1

        db.session.commit()
        flash(f'Importação concluída! {sucesso} funcionários importados com sucesso. {erros} linhas continham erros e foram ignoradas.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Ocorreu um erro ao processar o ficheiro: {e}', 'danger')

    return redirect(url_for('rh.listar_funcionarios'))

@rh.route('/funcionarios/template')
@login_required
def download_template():
    # (Este código não foi alterado)
    output = BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Modelo de Importação"

    headers = [
        "nome", "cpf", "email", "data_nascimento", "data_admissao", "salario",
        "cargo", "departamento", "rg", "sexo", "estado_civil", "telefone",
        "endereco", "cep", "cidade", "estado"
    ]
    sheet.append(headers)

    example = [
        "João da Silva", "123.456.789-00", "joao.silva@exemplo.com", "1990-05-15 00:00:00", "2025-08-01 00:00:00",
        "3500.50", "Analista Financeiro", "Financeiro", "1234567", "M", "Casado(a)",
        "(51) 99999-8888", "Rua das Flores, 123", "90000-000", "Porto Alegre", "RS"
    ]
    sheet.append(example)
    
    workbook.save(output)
    output.seek(0)
    
    return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment;filename=template_funcionarios.xlsx"})

@rh.route('/departamentos')
@login_required
@require_plan('medio')
def listar_departamentos():
    # (Este código não foi alterado)
    departamentos = Departamento.query.filter_by(empresa_id=current_user.empresa_id).order_by(Departamento.nome).all()
    return render_template('rh/departamentos.html', departamentos=departamentos)

@rh.route('/departamentos/novo', methods=['GET', 'POST'])
@login_required
@require_plan('medio')
def novo_departamento():
    # (Este código não foi alterado)
    if request.method == 'POST':
        novo_depto = Departamento(
            nome=request.form['nome'],
            descricao=request.form['descricao'],
            empresa_id=current_user.empresa_id
        )
        db.session.add(novo_depto)
        db.session.commit()
        flash('Departamento criado com sucesso!', 'success')
        return redirect(url_for('rh.listar_departamentos'))
    return render_template('rh/departamento_form.html', title="Novo Departamento")

@rh.route('/departamentos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@require_plan('medio')
def editar_departamento(id):
    # (Este código não foi alterado)
    depto = Departamento.query.get_or_404(id)
    if depto.empresa_id != current_user.empresa_id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('rh.listar_departamentos'))
    
    if request.method == 'POST':
        depto.nome = request.form['nome']
        depto.descricao = request.form['descricao']
        db.session.commit()
        flash('Departamento atualizado com sucesso!', 'success')
        return redirect(url_for('rh.listar_departamentos'))
    return render_template('rh/departamento_form.html', title="Editar Departamento", departamento=depto)

@rh.route('/cargos')
@login_required
@require_plan('medio')
def listar_cargos():
    # (Este código não foi alterado)
    cargos = Cargo.query.filter_by(empresa_id=current_user.empresa_id).order_by(Cargo.nome).all()
    return render_template('rh/cargos.html', cargos=cargos)

@rh.route('/cargos/novo', methods=['GET', 'POST'])
@login_required
@require_plan('medio')
def novo_cargo():
    # (Este código não foi alterado)
    if request.method == 'POST':
        salario_str = request.form['salario_base'].replace('R$', '').replace('.', '').replace(',', '.').strip()
        novo_cargo = Cargo(
            nome=request.form['nome'],
            cbo=request.form['cbo'],
            descricao=request.form['descricao'],
            salario_base=float(salario_str),
            nivel=request.form['nivel'],
            empresa_id=current_user.empresa_id
        )
        db.session.add(novo_cargo)
        db.session.commit()
        flash('Cargo criado com sucesso!', 'success')
        return redirect(url_for('rh.listar_cargos'))
    
    cbo_data = {}
    try:
        cbo_file_path = os.path.join(current_app.static_folder, 'data', 'cbo.json')
        with open(cbo_file_path, 'r', encoding='utf-8') as f:
            cbo_data = json.load(f)
    except Exception as e:
        flash(f"Atenção: Não foi possível carregar a lista de CBOs. A busca automática está desativada. Erro: {e}", "warning")

    return render_template('rh/cargo_form.html', title="Novo Cargo", cbo_data=cbo_data)

@rh.route('/cargos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@require_plan('medio')
def editar_cargo(id):
    # (Este código não foi alterado)
    cargo = Cargo.query.get_or_404(id)
    if cargo.empresa_id != current_user.empresa_id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('rh.listar_cargos'))

    if request.method == 'POST':
        salario_str = request.form['salario_base'].replace('R$', '').replace('.', '').replace(',', '.').strip()
        cargo.nome = request.form['nome']
        cargo.cbo = request.form['cbo']
        cargo.descricao = request.form['descricao']
        cargo.salario_base = float(salario_str)
        cargo.nivel = request.form['nivel']
        db.session.commit()
        flash('Cargo atualizado com sucesso!', 'success')
        return redirect(url_for('rh.listar_cargos'))
    
    cbo_data = {}
    try:
        cbo_file_path = os.path.join(current_app.static_folder, 'data', 'cbo.json')
        with open(cbo_file_path, 'r', encoding='utf-8') as f:
            cbo_data = json.load(f)
    except Exception as e:
        flash(f"Atenção: Não foi possível carregar a lista de CBOs. A busca automática está desativada. Erro: {e}", "warning")

    return render_template('rh/cargo_form.html', title="Editar Cargo", cargo=cargo, cbo_data=cbo_data)

@rh.route('/funcionarios/<int:id>/registrar_afastamento', methods=['GET', 'POST'])
@login_required
@require_plan('medio')
def registrar_afastamento(id):
    # (Este código não foi alterado)
    funcionario = Funcionario.query.get_or_404(id)
    if funcionario.empresa_id != current_user.empresa_id:
        flash('Acesso negado.', 'danger')
        return redirect(url_for('rh.listar_funcionarios'))

    if request.method == 'POST':
        try:
            novo_afastamento = Afastamento(
                funcionario_id=id,
                motivo=request.form.get('motivo'),
                data_inicio=datetime.strptime(request.form.get('data_inicio'), '%Y-%m-%d').date(),
                data_fim=datetime.strptime(request.form.get('data_fim'), '%Y-%m-%d').date(),
                observacoes=request.form.get('observacoes')
            )
            db.session.add(novo_afastamento)
            db.session.commit()
            flash('Registro de afastamento adicionado com sucesso!', 'success')
            return redirect(url_for('rh.ver_funcionario', id=id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao registrar afastamento: {str(e)}', 'danger')

    return render_template('rh/afastamento_form.html', funcionario=funcionario)